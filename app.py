from openpyxl import load_workbook

wb = load_workbook(filename = './NF-SA Template 160519.xlsx')
sheet_ranges = wb['SA-Ratios']
letters = tuple(map(chr, range(65, 91)))
digits = ('0', '1', '2', '3', '4', '5', '6', '7', '8', '9')
cell_value_dict = {}

#Returns value or formula of a cell in a different sheet.
def getValueFromSheet(cell, skip_places):    
    sheet_name_cell = cell.partition('!')
    sheet_name = sheet_name_cell[0].replace("'",'')
    sheet_ranges_temp = wb[sheet_name]
    cell_loc = sheet_name_cell[2]
    new_cell_loc = letters[letters.index(cell_loc[:1]) + skip_places] + cell_loc[1:]
    cell_value = sheet_ranges_temp[new_cell_loc].value
    # print(sheet_name, cell_value, skip_places)
    return str(cell_value).strip()

#Returns the value or formula of a cell. 
def getValue(cell, skip_places, rhs_char_list):
    value = cell
    value_cell_loc = ''
    if len(cell) > 0:
        if cell[:1] == "'":
            value = getValueFromSheet(cell, skip_places)
        elif cell[:1] in letters and cell[-1] in digits:
            value_cell_loc = letters[letters.index(cell[:1]) + skip_places] + cell[1:]
            value = sheet_ranges[str(value_cell_loc)].value
        else:
            value = cell
    return value
        
#Takes an excel formula as input.
#Returns the value of the parsed formula in english according to thier respective cell locations.
def parseFormula(formula, rhs_char_list, skip_places=None):
    if formula[:1] != '=':
        formula_list = []
        formula_list.append(formula)
        return formula_list
    else:
        formula = formula.replace('$','').replace('=', '').strip()
        index = 0
        cell = ''
        cell_value = ''
        skip = False
        for index, char in enumerate(formula):  #Loop over each character to look for delemiters.
            if (char in ['+','-','/','*','(',')',':','&']) and (not skip):
                if cell in cell_value_dict.keys():
                    rhs_char_list.append(cell_value_dict[cell])
                else:
                    cell_value = str(getValue(cell, skip_places,rhs_char_list))
                    if cell_value != None and cell_value[:1] == '=':
                        cell_value = parseFormula(cell_value, rhs_char_list, 0)
                        cell_value = cell_value[0]
                    else:
                        rhs_char_list.append(cell_value)
                        cell_value_dict[cell] = cell_value
                        # print(cell.strip(), cell_value.strip())
                rhs_char_list.append(char)                               
                cell = ''
            elif len(formula) == index+1:
                cell += char
                if cell in cell_value_dict.keys():
                    rhs_char_list.append(cell_value_dict[cell])
                else:
                    cell_value = getValue(cell, skip_places, rhs_char_list) 
                    if cell_value != None and cell_value[:1] == '=':
                        cell_value = parseFormula(cell_value, rhs_char_list, 0)
                        cell_value = cell_value[0]
                    else:
                        rhs_char_list.append(cell_value)
                        cell_value_dict[cell] = cell_value
            elif char in["'"]:
                cell += char
                skip = not skip
            else:
                cell += char 
        # print(rhs_char_list)
        return rhs_char_list

file = open('./output.txt','w') 
#Loop over every cell Bx to look for the cells which have a value or a formula in cell Cx.   
for row in sheet_ranges.iter_rows( min_col=2, max_col=2):
    for index,cell in enumerate(row):
        lhs_cell_loc = cell.coordinate
        rhs_cell_loc = letters[letters.index(lhs_cell_loc[:1])+1] + lhs_cell_loc[1:]
        if cell.value and sheet_ranges[rhs_cell_loc].value:
            formula_text = ''
            lhs = cell.value
            rhs = sheet_ranges[rhs_cell_loc].value
            lhs_char_list = []
            rhs_char_list = []
            lhs = parseFormula(str(lhs), lhs_char_list, 0)
            lhs_string = ''.join(word for word in lhs)
            rhs = parseFormula(str(rhs), rhs_char_list, -1)
            rhs_string = ''.join(word for word in rhs)
            formula = lhs_string + ' = ' + rhs_string
            file.write(formula+'\n')
            print(cell, formula)
file.close() 
